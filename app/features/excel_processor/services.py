import logging
import pandas as pd
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import Dict, Any, List

@dataclass
class ExcelDocument:
    filename: str
    summary: Dict[str, Any]

class ExcelProcessorService:
    def process_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Process an Excel file and extract useful information.
        
        Args:
            file_path: Path to the uploaded Excel file
            
        Returns:
            Dictionary containing summary information about the Excel file
        """
        logging.debug(f"Processing file: {file_path}")
        if file_path.endswith('.csv'):
            # Read CSV file
            logging.debug("Reading CSV file.")
            df = pd.read_csv(file_path)
            summary = self._generate_summary_from_dataframe(df, 'CSV')
        else:
            # Load the workbook to get sheet names and metadata
            logging.debug("Loading Excel workbook.")
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            
            # Get basic information
            summary = {
                'sheet_names': workbook.sheetnames,
                'active_sheet': workbook.active.title,
                'sheet_summaries': {}
            }
            
            # Use pandas to read each sheet and get statistics
            for sheet_name in workbook.sheetnames:
                # Read sheet with pandas
                logging.debug(f"Reading sheet: {sheet_name}")
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_summary = self._generate_summary_from_dataframe(df, sheet_name)
                summary['sheet_summaries'][sheet_name] = sheet_summary
        
        logging.debug(f"Generated summary before return: {summary}")
        # Generate a text summary without using AI
        text_summary = self._generate_simple_text_summary(summary)
        logging.debug(f"Generated text summary: {text_summary}")
        return {'summary': summary, 'text_summary': text_summary}

    def _generate_simple_text_summary(self, summary: Dict[str, Any]) -> str:
        """Generate a simple text summary without using external APIs."""
        if 'sheet_names' in summary:
            # For Excel files with multiple sheets
            total_sheets = len(summary['sheet_names'])
            total_rows = sum(summary['sheet_summaries'][sheet].get('rows', 0) for sheet in summary['sheet_names'])
            
            text_parts = [
                f"Excel file contains {total_sheets} sheet(s): {', '.join(summary['sheet_names'])}.",
                f"Total of {total_rows} rows across all sheets.",
            ]
            
            # Add info about each sheet
            for sheet_name in summary['sheet_names']:
                sheet_data = summary['sheet_summaries'][sheet_name]
                text_parts.append(
                    f"\nSheet '{sheet_name}': {sheet_data['rows']} rows, {sheet_data['columns']} columns."
                )
                
                # Add column names
                if sheet_data.get('column_names'):
                    text_parts.append(f"Columns: {', '.join(sheet_data['column_names'])}.")
                
                # Add numeric column stats if available
                if sheet_data.get('numeric_columns') and sheet_data.get('statistics'):
                    num_cols = len(sheet_data['numeric_columns'])
                    text_parts.append(f"Contains {num_cols} numeric column(s): {', '.join(sheet_data['numeric_columns'])}.")
                    
                    # Add key statistics for numeric columns
                    for col in sheet_data['numeric_columns']:
                        if col in sheet_data['statistics']:
                            stats = sheet_data['statistics'][col]
                            text_parts.append(
                                f"  • {col}: min={stats['min']:.1f}, max={stats['max']:.1f}, avg={stats['mean']:.1f}"
                            )
            
            return "\n".join(text_parts)
        else:
            # For CSV files
            rows = summary.get('rows', 0)
            cols = summary.get('columns', 0)
            col_names = summary.get('column_names', [])
            num_cols = summary.get('numeric_columns', [])
            
            text_parts = [
                f"CSV file with {rows} rows and {cols} columns.",
                f"Columns: {', '.join(col_names)}."
            ]
            
            if num_cols:
                text_parts.append(f"Contains {len(num_cols)} numeric column(s): {', '.join(num_cols)}.")
                
                # Add key statistics for numeric columns
                for col in num_cols:
                    if 'statistics' in summary and col in summary['statistics']:
                        stats = summary['statistics'][col]
                        text_parts.append(
                            f"  • {col}: min={stats['min']:.1f}, max={stats['max']:.1f}, avg={stats['mean']:.1f}"
                        )
            
            return "\n".join(text_parts)

    def _generate_summary_from_dataframe(self, df, sheet_name):
        """Generate a summary from a DataFrame."""
        sheet_summary = {
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'numeric_columns': df.select_dtypes(include=['number']).columns.tolist(),
            'non_null_counts': df.count().to_dict(),
        }
        
        # Add basic statistics for numeric columns
        if len(sheet_summary['numeric_columns']) > 0:
            stats = df[sheet_summary['numeric_columns']].describe().to_dict()
            sheet_summary['statistics'] = stats
        
        return sheet_summary
